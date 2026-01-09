"""
Bid Estimator - Expert civil engineering bid quantity and cost estimation
Calculates material, labor, and equipment costs from bid documents
"""

import os
import io
import json
import re
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field

import fitz  # PyMuPDF
from openai import OpenAI
import openpyxl


@dataclass
class LineItemEstimate:
    """A single line item estimate with material, labor, equipment breakdown"""
    item_number: str = ""
    description: str = ""
    quantity: float = 0.0
    unit: str = ""
    
    # Cost breakdown
    material_unit_cost: float = 0.0
    labor_unit_cost: float = 0.0
    equipment_unit_cost: float = 0.0
    unit_price: float = 0.0
    total_price: float = 0.0
    
    # Details
    category: str = ""
    crew_size: str = ""
    production_rate: str = ""
    material_specs: str = ""
    notes: str = ""
    confidence: float = 0.0


class BidEstimator:
    """
    Expert civil engineering bid estimator.
    Analyzes bid documents and calculates material, labor, and equipment costs.
    """
    
    # Civil engineering estimation prompt
    ESTIMATION_PROMPT = """You are an expert civil engineering estimator with 20+ years of experience bidding on municipal infrastructure, road, utility, and site development projects.

Analyze the following bid document content and provide detailed line item estimates.

For EACH line item or work item identified, provide:

1. QUANTITIES:
   - Item description (use standard pay item descriptions)
   - Estimated quantity
   - Unit of measure (CY, LF, SY, TON, EA, LS, etc.)

2. MATERIAL COSTS:
   - Specific materials required
   - Material unit cost estimate
   - Material specifications if noted

3. LABOR COSTS:
   - Recommended crew size and composition
   - Estimated production rate (units per day)
   - Labor unit cost estimate
   - Special skill requirements

4. EQUIPMENT COSTS:
   - Required equipment
   - Equipment unit cost estimate
   - Mobilization considerations

5. UNIT PRICE:
   - Total unit price (material + labor + equipment + overhead/profit)
   - Assume 15-20% overhead and profit markup

IMPORTANT CONSIDERATIONS:
- Account for local prevailing wages if noted
- Consider mobilization/demobilization
- Note any special conditions (night work, traffic control, phasing)
- Identify permit requirements
- Flag any unusual specifications or risks
- Consider weather/seasonal factors

Return a JSON object:
{
    "project_summary": {
        "project_name": "",
        "location": "",
        "owner": "",
        "bid_date": "",
        "project_type": "road/utility/site/mixed",
        "special_conditions": [],
        "key_risks": []
    },
    "line_items": [
        {
            "item_number": "",
            "description": "",
            "quantity": 0,
            "unit": "",
            "category": "earthwork/paving/utilities/structures/traffic/erosion/landscape/general",
            "material": {
                "description": "",
                "unit_cost": 0,
                "specs": ""
            },
            "labor": {
                "crew": "",
                "production_rate": "",
                "unit_cost": 0,
                "notes": ""
            },
            "equipment": {
                "required": [],
                "unit_cost": 0
            },
            "unit_price": 0,
            "total_price": 0,
            "confidence": 0.0-1.0,
            "notes": ""
        }
    ],
    "summary_by_category": {
        "earthwork": {"subtotal": 0, "items": 0},
        "paving": {"subtotal": 0, "items": 0},
        "utilities": {"subtotal": 0, "items": 0},
        "structures": {"subtotal": 0, "items": 0},
        "traffic": {"subtotal": 0, "items": 0},
        "erosion": {"subtotal": 0, "items": 0},
        "general": {"subtotal": 0, "items": 0}
    },
    "bid_total": {
        "subtotal": 0,
        "contingency_pct": 10,
        "contingency_amt": 0,
        "total": 0
    },
    "estimator_notes": [
        "Key observations and recommendations"
    ]
}

Only return the JSON object, no other text."""

    PROPOSAL_REVIEW_PROMPT = """You are an expert civil engineering estimator reviewing a bid proposal.

Compare this proposal against typical industry standards and best practices.

Analyze:
1. Are quantities reasonable for the scope described?
2. Are unit prices competitive but realistic?
3. Is the material/labor/equipment breakdown appropriate?
4. Are there any missing items that should be included?
5. Are there any items that seem overpriced or underpriced?
6. What risks or issues should be addressed?

Provide specific, actionable feedback to improve the bid.

Return a JSON object:
{
    "overall_assessment": "strong/adequate/needs_work/weak",
    "competitiveness_score": 0-100,
    "line_item_review": [
        {
            "item": "",
            "status": "good/high/low/missing",
            "current_price": 0,
            "recommended_price": 0,
            "variance_pct": 0,
            "comments": ""
        }
    ],
    "missing_items": [
        {"description": "", "estimated_cost": 0, "reason": ""}
    ],
    "risks": [
        {"risk": "", "severity": "high/medium/low", "mitigation": ""}
    ],
    "recommendations": [
        "Specific recommendations to improve the bid"
    ],
    "summary": "Executive summary of the proposal review"
}

Only return the JSON object, no other text."""

    def __init__(self, api_key: Optional[str] = None):
        """Initialize the bid estimator with OpenAI API key."""
        self.api_key = api_key or os.environ.get('OPENAI_API_KEY')
        if not self.api_key:
            raise ValueError("OpenAI API key is required")
        self.client = OpenAI(api_key=self.api_key)
        self.model = "gpt-4o"
    
    def extract_text_from_pdf(self, pdf_path: str) -> str:
        """Extract all text content from a PDF."""
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()
        return text
    
    def extract_text_from_excel(self, excel_path: str) -> str:
        """Extract all content from Excel as text."""
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        text = ""
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n=== SHEET: {sheet_name} ===\n"
            
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip(" |"):
                    text += row_text + "\n"
        
        wb.close()
        return text
    
    def analyze_bid_documents(self, file_paths: List[str]) -> Dict[str, Any]:
        """
        Analyze bid documents and generate detailed estimates.
        
        Args:
            file_paths: List of paths to bid documents (PDF, Excel)
            
        Returns:
            Complete bid analysis with line item estimates
        """
        # Extract text from all documents
        combined_text = ""
        for path in file_paths:
            ext = os.path.splitext(path)[1].lower()
            if ext == '.pdf':
                combined_text += f"\n--- Document: {os.path.basename(path)} ---\n"
                combined_text += self.extract_text_from_pdf(path)
            elif ext in ['.xlsx', '.xls', '.xlsm']:
                combined_text += f"\n--- Document: {os.path.basename(path)} ---\n"
                combined_text += self.extract_text_from_excel(path)
        
        # Limit text for API
        if len(combined_text) > 50000:
            combined_text = combined_text[:50000] + "\n[... truncated ...]"
        
        # Call GPT for analysis
        response = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert civil engineering estimator. Provide detailed, accurate bid estimates based on current market rates and industry standards."
                },
                {
                    "role": "user",
                    "content": f"{self.ESTIMATION_PROMPT}\n\nBID DOCUMENT CONTENT:\n{combined_text}"
                }
            ],
            max_tokens=8000,
            temperature=0.3
        )
        
        return self._parse_response(response.choices[0].message.content)
    
    def review_proposal(self, proposal_paths: List[str], bid_doc_paths: Optional[List[str]] = None) -> Dict[str, Any]:
        """
        Review an existing proposal and provide feedback.
        
        Args:
            proposal_paths: Paths to the proposal being reviewed
            bid_doc_paths: Optional paths to original bid documents for comparison
            
        Returns:
            Proposal review with recommendations
        """
        # Extract proposal content
        proposal_text = ""
        for path in proposal_paths:
            ext = os.path.splitext(path)[1].lower()
            if ext == '.pdf':
                proposal_text += self.extract_text_from_pdf(path)
            elif ext in ['.xlsx', '.xls', '.xlsm']:
                proposal_text += self.extract_text_from_excel(path)
        
        # Extract bid doc content if provided
        bid_doc_text = ""
        if bid_doc_paths:
            for path in bid_doc_paths:
                ext = os.path.splitext(path)[1].lower()
                if ext == '.pdf':
                    bid_doc_text += self.extract_text_from_pdf(path)
                elif ext in ['.xlsx', '.xls', '.xlsm']:
                    bid_doc_text += self.extract_text_from_excel(path)
        
        context = f"PROPOSAL CONTENT:\n{proposal_text[:30000]}"
        if bid_doc_text:
            context += f"\n\nORIGINAL BID DOCUMENTS:\n{bid_doc_text[:20000]}"
        
        response = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert civil engineering estimator reviewing bid proposals. Provide specific, actionable feedback."
                },
                {
                    "role": "user",
                    "content": f"{self.PROPOSAL_REVIEW_PROMPT}\n\n{context}"
                }
            ],
            max_tokens=6000,
            temperature=0.3
        )
        
        return self._parse_response(response.choices[0].message.content)
    
    def generate_estimate_from_scope(self, scope_description: str) -> Dict[str, Any]:
        """
        Generate a preliminary estimate from a scope description.
        
        Args:
            scope_description: Text description of the project scope
            
        Returns:
            Preliminary estimate with line items
        """
        response = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert civil engineering estimator. Generate detailed preliminary estimates."
                },
                {
                    "role": "user",
                    "content": f"{self.ESTIMATION_PROMPT}\n\nPROJECT SCOPE:\n{scope_description}"
                }
            ],
            max_tokens=6000,
            temperature=0.3
        )
        
        return self._parse_response(response.choices[0].message.content)
    
    def _parse_response(self, content: str) -> Dict[str, Any]:
        """Parse GPT response to JSON."""
        try:
            # Clean up markdown code blocks
            if content.startswith('```'):
                content = content.split('```')[1]
                if content.startswith('json'):
                    content = content[4:]
            if content.endswith('```'):
                content = content[:-3]
            
            return json.loads(content.strip())
            
        except json.JSONDecodeError:
            # Try to extract JSON
            start = content.find('{')
            end = content.rfind('}') + 1
            if start != -1 and end > start:
                try:
                    return json.loads(content[start:end])
                except:
                    pass
            
            return {
                'error': 'Failed to parse response',
                'raw_content': content[:2000]
            }
    
    def calculate_totals(self, line_items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate totals from line items."""
        totals = {
            'material_total': 0,
            'labor_total': 0,
            'equipment_total': 0,
            'subtotal': 0,
            'by_category': {}
        }
        
        for item in line_items:
            qty = float(item.get('quantity', 0) or 0)
            mat = float(item.get('material', {}).get('unit_cost', 0) or 0) * qty
            lab = float(item.get('labor', {}).get('unit_cost', 0) or 0) * qty
            equip = float(item.get('equipment', {}).get('unit_cost', 0) or 0) * qty
            total = float(item.get('total_price', 0) or 0)
            
            totals['material_total'] += mat
            totals['labor_total'] += lab
            totals['equipment_total'] += equip
            totals['subtotal'] += total
            
            cat = item.get('category', 'general')
            if cat not in totals['by_category']:
                totals['by_category'][cat] = {'subtotal': 0, 'items': 0}
            totals['by_category'][cat]['subtotal'] += total
            totals['by_category'][cat]['items'] += 1
        
        return totals
    
    def format_currency(self, amount: float) -> str:
        """Format number as currency."""
        return f"${amount:,.2f}"
    
    def export_to_dict(self, estimate: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Convert estimate to flat list for export."""
        items = []
        for item in estimate.get('line_items', []):
            items.append({
                'Item No.': item.get('item_number', ''),
                'Description': item.get('description', ''),
                'Quantity': item.get('quantity', 0),
                'Unit': item.get('unit', ''),
                'Material Cost': item.get('material', {}).get('unit_cost', 0),
                'Labor Cost': item.get('labor', {}).get('unit_cost', 0),
                'Equipment Cost': item.get('equipment', {}).get('unit_cost', 0),
                'Unit Price': item.get('unit_price', 0),
                'Total Price': item.get('total_price', 0),
                'Category': item.get('category', ''),
                'Notes': item.get('notes', '')
            })
        return items
