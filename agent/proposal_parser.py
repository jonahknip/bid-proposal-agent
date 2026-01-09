"""
Proposal Parser - Extract requirements, line items, and scope from RFP/bid documents
Supports PDF and Excel formats
"""

import os
import io
import json
import re
import base64
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass, field, asdict

import fitz  # PyMuPDF
from PIL import Image
from openai import OpenAI
import openpyxl


@dataclass
class BidLineItem:
    """A single line item from bid documents"""
    item_number: str = ""
    description: str = ""
    quantity: float = 0.0
    unit: str = ""
    unit_price: float = 0.0
    total_price: float = 0.0
    category: str = ""
    notes: str = ""
    required: bool = True


@dataclass
class ProposalRequirements:
    """Extracted requirements from proposal/RFP documents"""
    project_name: str = ""
    project_number: str = ""
    owner: str = ""
    location: str = ""
    bid_date: str = ""
    submission_deadline: str = ""
    scope_summary: str = ""
    line_items: List[BidLineItem] = field(default_factory=list)
    special_requirements: List[str] = field(default_factory=list)
    qualifications: List[str] = field(default_factory=list)
    bonding_requirements: str = ""
    insurance_requirements: str = ""
    liquidated_damages: str = ""
    completion_time: str = ""
    addenda: List[str] = field(default_factory=list)
    contacts: List[Dict[str, str]] = field(default_factory=list)
    documents_required: List[str] = field(default_factory=list)


class ProposalParser:
    """
    Parses RFP/bid documents to extract requirements, line items, and scope.
    Supports PDF and Excel formats.
    """
    
    # RFP parsing prompt for GPT-4
    RFP_EXTRACTION_PROMPT = """You are an expert at analyzing civil engineering bid documents and RFPs.
Analyze this document and extract all relevant bid information.

Extract the following information:

1. PROJECT IDENTIFICATION:
   - Project name
   - Project number
   - Owner/Client
   - Location
   - Engineer of Record

2. BID SCHEDULE:
   - Bid date/deadline
   - Submission requirements
   - Pre-bid meeting date (if any)

3. SCOPE OF WORK:
   - Brief summary of work
   - Major work items
   - Project phases (if any)

4. LINE ITEMS / BID SCHEDULE:
   Extract ALL line items from any bid schedule, pay item list, or quantity table.
   For each item:
   - Item number
   - Description
   - Estimated quantity
   - Unit of measure
   - Category (earthwork, paving, utilities, etc.)

5. SPECIAL REQUIREMENTS:
   - Qualifications required
   - Bonding requirements
   - Insurance requirements
   - Liquidated damages
   - Completion time/schedule
   - DBE/MBE requirements
   - Special provisions

6. DOCUMENTS REQUIRED:
   - List all documents required for bid submission

Return a JSON object:
{
    "project_info": {
        "project_name": "",
        "project_number": "",
        "owner": "",
        "location": "",
        "engineer": ""
    },
    "bid_schedule": {
        "bid_date": "",
        "submission_deadline": "",
        "pre_bid_meeting": ""
    },
    "scope_summary": "Brief description of the project scope",
    "line_items": [
        {
            "item_number": "",
            "description": "",
            "quantity": 0,
            "unit": "",
            "category": "",
            "notes": ""
        }
    ],
    "requirements": {
        "qualifications": [],
        "bonding": "",
        "insurance": "",
        "liquidated_damages": "",
        "completion_time": "",
        "special": []
    },
    "documents_required": [],
    "contacts": [
        {"name": "", "title": "", "email": "", "phone": ""}
    ]
}

Only return the JSON object, no other text."""

    EXCEL_BID_SCHEDULE_PROMPT = """You are analyzing a civil engineering bid schedule spreadsheet.
The data below is extracted from an Excel file. Parse it and extract all bid line items.

For each line item, identify:
- Item number
- Description  
- Quantity
- Unit of measure
- Unit price (if provided)
- Category (earthwork, paving, utilities, structures, erosion control, traffic control, landscaping, misc)

Return a JSON object:
{
    "line_items": [
        {
            "item_number": "",
            "description": "",
            "quantity": 0,
            "unit": "",
            "unit_price": 0,
            "category": "",
            "notes": ""
        }
    ],
    "totals": {
        "subtotal": 0,
        "contingency": 0,
        "total": 0
    },
    "notes": "Any relevant notes about the bid schedule"
}

Only return the JSON object, no other text.

EXCEL DATA:
"""

    def __init__(self, api_key: Optional[str] = None):
        """Initialize the proposal parser with OpenAI API key."""
        self.api_key = api_key or os.environ.get('OPENAI_API_KEY')
        if not self.api_key:
            raise ValueError("OpenAI API key is required")
        self.client = OpenAI(api_key=self.api_key)
        self.model = "gpt-4o"
    
    def parse_pdf(self, pdf_path: str, max_pages: int = 20) -> Dict[str, Any]:
        """
        Parse an RFP/bid document PDF.
        
        Args:
            pdf_path: Path to the PDF file
            max_pages: Maximum pages to analyze with vision
            
        Returns:
            Extracted proposal requirements
        """
        doc = fitz.open(pdf_path)
        
        # Extract all text first
        full_text = ""
        for page in doc:
            full_text += page.get_text() + "\n"
        
        # Also get images of key pages for vision analysis
        images = []
        for page_num in range(min(len(doc), max_pages)):
            page = doc[page_num]
            zoom = 150 / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            
            # Resize if needed
            max_dim = 1536
            if max(img.size) > max_dim:
                ratio = max_dim / max(img.size)
                new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
                img = img.resize(new_size, Image.LANCZOS)
            
            buffer = io.BytesIO()
            img.save(buffer, format="PNG")
            base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')
            images.append(base64_image)
        
        doc.close()
        
        # Build message content with text and images
        content = [
            {"type": "text", "text": f"{self.RFP_EXTRACTION_PROMPT}\n\nDOCUMENT TEXT:\n{full_text[:15000]}"}
        ]
        
        # Add images (limit to 5 for token efficiency)
        for img_b64 in images[:5]:
            content.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{img_b64}",
                    "detail": "high"
                }
            })
        
        # Call GPT-4 Vision
        response = self.client.chat.completions.create(
            model=self.model,
            messages=[{"role": "user", "content": content}],
            max_tokens=4000,
            temperature=0.2
        )
        
        try:
            result_text = response.choices[0].message.content
            # Clean up response
            if result_text.startswith('```'):
                result_text = result_text.split('```')[1]
                if result_text.startswith('json'):
                    result_text = result_text[4:]
            if result_text.endswith('```'):
                result_text = result_text[:-3]
            
            result = json.loads(result_text.strip())
            result['source_file'] = os.path.basename(pdf_path)
            result['pages_analyzed'] = min(len(images), max_pages)
            return result
            
        except json.JSONDecodeError:
            # Try to extract JSON
            content_text = response.choices[0].message.content
            start = content_text.find('{')
            end = content_text.rfind('}') + 1
            if start != -1 and end > start:
                try:
                    result = json.loads(content_text[start:end])
                    result['source_file'] = os.path.basename(pdf_path)
                    return result
                except:
                    pass
            
            return {
                'error': 'Failed to parse document',
                'source_file': os.path.basename(pdf_path),
                'raw_text': full_text[:5000]
            }
    
    def parse_excel(self, excel_path: str) -> Dict[str, Any]:
        """
        Parse an Excel bid schedule.
        
        Args:
            excel_path: Path to the Excel file
            
        Returns:
            Extracted line items and totals
        """
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        all_data = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {
                'sheet_name': sheet_name,
                'rows': []
            }
            
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    sheet_data['rows'].append([str(cell) if cell is not None else '' for cell in row])
            
            all_data.append(sheet_data)
        
        wb.close()
        
        # Convert to text for GPT analysis
        text_repr = ""
        for sheet in all_data:
            text_repr += f"\n=== SHEET: {sheet['sheet_name']} ===\n"
            for row in sheet['rows'][:100]:  # Limit rows
                text_repr += " | ".join(row) + "\n"
        
        # Use GPT to parse the structure
        response = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {"role": "user", "content": f"{self.EXCEL_BID_SCHEDULE_PROMPT}\n{text_repr}"}
            ],
            max_tokens=4000,
            temperature=0.2
        )
        
        try:
            result_text = response.choices[0].message.content
            if result_text.startswith('```'):
                result_text = result_text.split('```')[1]
                if result_text.startswith('json'):
                    result_text = result_text[4:]
            if result_text.endswith('```'):
                result_text = result_text[:-3]
            
            result = json.loads(result_text.strip())
            result['source_file'] = os.path.basename(excel_path)
            result['sheets_found'] = [s['sheet_name'] for s in all_data]
            return result
            
        except json.JSONDecodeError:
            content_text = response.choices[0].message.content
            start = content_text.find('{')
            end = content_text.rfind('}') + 1
            if start != -1 and end > start:
                try:
                    result = json.loads(content_text[start:end])
                    result['source_file'] = os.path.basename(excel_path)
                    return result
                except:
                    pass
            
            return {
                'error': 'Failed to parse Excel file',
                'source_file': os.path.basename(excel_path),
                'sheets_found': [s['sheet_name'] for s in all_data]
            }
    
    def parse_bid_document(self, file_path: str) -> Dict[str, Any]:
        """
        Parse a bid document (auto-detects PDF or Excel).
        
        Args:
            file_path: Path to the document
            
        Returns:
            Extracted requirements and line items
        """
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.pdf':
            return self.parse_pdf(file_path)
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            return self.parse_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")
    
    def parse_multiple_documents(self, file_paths: List[str]) -> Dict[str, Any]:
        """
        Parse multiple bid documents and combine results.
        
        Args:
            file_paths: List of paths to documents
            
        Returns:
            Combined extracted requirements
        """
        results = []
        all_line_items = []
        
        for path in file_paths:
            try:
                result = self.parse_bid_document(path)
                results.append(result)
                
                # Collect line items
                if 'line_items' in result:
                    for item in result['line_items']:
                        item['source_file'] = os.path.basename(path)
                        all_line_items.append(item)
                        
            except Exception as e:
                results.append({
                    'error': str(e),
                    'source_file': os.path.basename(path)
                })
        
        # Combine project info from first successful parse
        combined_project_info = {}
        for r in results:
            if 'project_info' in r:
                combined_project_info = r['project_info']
                break
        
        # Combine requirements
        combined_requirements = {
            'qualifications': [],
            'bonding': '',
            'insurance': '',
            'special': []
        }
        for r in results:
            if 'requirements' in r:
                req = r['requirements']
                combined_requirements['qualifications'].extend(req.get('qualifications', []))
                if req.get('bonding'):
                    combined_requirements['bonding'] = req['bonding']
                if req.get('insurance'):
                    combined_requirements['insurance'] = req['insurance']
                combined_requirements['special'].extend(req.get('special', []))
        
        return {
            'project_info': combined_project_info,
            'individual_results': results,
            'combined_line_items': all_line_items,
            'requirements': combined_requirements,
            'files_processed': len(file_paths)
        }
    
    def extract_line_items_table(self, result: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Extract just the line items from parsed results in a clean table format.
        
        Args:
            result: Parsed proposal results
            
        Returns:
            List of line items for tabular display
        """
        items = result.get('line_items', []) or result.get('combined_line_items', [])
        
        clean_items = []
        for i, item in enumerate(items, 1):
            clean_items.append({
                'No.': item.get('item_number', str(i)),
                'Description': item.get('description', ''),
                'Quantity': item.get('quantity', 0),
                'Unit': item.get('unit', ''),
                'Unit Price': item.get('unit_price', 0),
                'Category': item.get('category', ''),
                'Source': item.get('source_file', '')
            })
        
        return clean_items
    
    def generate_requirements_checklist(self, result: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        Generate a checklist of requirements from parsed proposal.
        
        Args:
            result: Parsed proposal results
            
        Returns:
            List of requirement items for checklist
        """
        checklist = []
        
        # Project info requirements
        project = result.get('project_info', {})
        if project.get('project_name'):
            checklist.append({
                'category': 'Project',
                'requirement': f"Project: {project.get('project_name')}",
                'status': 'info'
            })
        
        # Bid schedule
        bid_schedule = result.get('bid_schedule', {})
        if bid_schedule.get('bid_date'):
            checklist.append({
                'category': 'Deadline',
                'requirement': f"Bid Due: {bid_schedule.get('bid_date')}",
                'status': 'critical'
            })
        
        # Requirements
        requirements = result.get('requirements', {})
        
        if requirements.get('bonding'):
            checklist.append({
                'category': 'Bonding',
                'requirement': requirements.get('bonding'),
                'status': 'required'
            })
        
        if requirements.get('insurance'):
            checklist.append({
                'category': 'Insurance',
                'requirement': requirements.get('insurance'),
                'status': 'required'
            })
        
        for qual in requirements.get('qualifications', []):
            checklist.append({
                'category': 'Qualification',
                'requirement': qual,
                'status': 'required'
            })
        
        for spec in requirements.get('special', []):
            checklist.append({
                'category': 'Special',
                'requirement': spec,
                'status': 'required'
            })
        
        # Documents required
        for doc in result.get('documents_required', []):
            checklist.append({
                'category': 'Document',
                'requirement': doc,
                'status': 'required'
            })
        
        return checklist
    
    def categorize_line_items(self, items: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Organize line items by category.
        
        Args:
            items: List of line items
            
        Returns:
            Dictionary of items organized by category
        """
        categories = {
            'earthwork': [],
            'paving': [],
            'utilities': [],
            'structures': [],
            'erosion_control': [],
            'traffic_control': [],
            'landscaping': [],
            'survey': [],
            'misc': []
        }
        
        for item in items:
            cat = item.get('category', 'misc').lower().replace(' ', '_')
            if cat not in categories:
                cat = 'misc'
            categories[cat].append(item)
        
        # Remove empty categories
        return {k: v for k, v in categories.items() if v}
